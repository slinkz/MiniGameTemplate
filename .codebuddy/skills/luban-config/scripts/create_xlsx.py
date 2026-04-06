#!/usr/bin/env python3
"""
create_xlsx.py — 创建符合 Luban v4.x 规范的 xlsx 数据文件。

Usage:
    python create_xlsx.py --output <path.xlsx> --sheet <SheetName> --fields "name:type:comment,..."

Example:
    python create_xlsx.py \
        --output UnityProj/DataTables/Datas/mydata.xlsx \
        --sheet TbMyData \
        --fields "id:int:ID,name:string:名称,desc:string:描述,price:int:价格"

Requires: openpyxl (pip install openpyxl)
"""

import argparse
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# --- Styles ---
HEADER_FONT = Font(bold=True, size=10, name="Arial")
DATA_FONT = Font(size=10, name="Arial")
META_FILL = PatternFill("solid", fgColor="E2EFDA")       # 浅绿 — ##var / ##type 行
COMMENT_FILL = PatternFill("solid", fgColor="D9E2F3")    # 浅蓝 — ## 注释行
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_row(ws, row_num, font, fill=None):
    """Apply font, border, alignment (and optional fill) to all cells in a row."""
    for cell in ws[row_num]:
        cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill


def parse_fields(fields_str: str):
    """Parse 'name:type:comment,...' into list of (name, type, comment)."""
    fields = []
    for part in fields_str.split(","):
        part = part.strip()
        if not part:
            continue
        tokens = part.split(":")
        if len(tokens) < 2:
            print(f"WARNING: Skipping malformed field '{part}' (expected name:type[:comment])", file=sys.stderr)
            continue
        name = tokens[0].strip()
        ftype = tokens[1].strip()
        comment = tokens[2].strip() if len(tokens) >= 3 else name
        fields.append((name, ftype, comment))
    return fields


def create_xlsx(output_path: str, sheet_name: str, fields: list):
    """Create a Luban-compliant xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Row 1: ##var + field names
    var_row = ["##var"] + [f[0] for f in fields]
    ws.append(var_row)

    # Row 2: ##type + field types
    type_row = ["##type"] + [f[1] for f in fields]
    ws.append(type_row)

    # Row 3: ## + comments (Chinese headers)
    comment_row = ["##"] + [f[2] for f in fields]
    ws.append(comment_row)

    # Apply styles
    style_row(ws, 1, HEADER_FONT, META_FILL)
    style_row(ws, 2, HEADER_FONT, META_FILL)
    style_row(ws, 3, HEADER_FONT, COMMENT_FILL)

    # Auto-width columns
    ws.column_dimensions["A"].width = 10
    for i, (name, ftype, comment) in enumerate(fields, start=2):
        col_letter = chr(64 + i) if i <= 26 else None
        if col_letter:
            # Width = max of name/comment length * 2 (rough CJK estimate), min 12
            width = max(len(name) * 1.5, len(comment) * 2.2, 12)
            ws.column_dimensions[col_letter].width = min(width, 30)

    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    print(f"OK: Created {output_path}")
    print(f"    Sheet: {sheet_name}")
    print(f"    Fields: {len(fields)}")
    for name, ftype, comment in fields:
        print(f"      - {name} ({ftype}) [{comment}]")


def main():
    parser = argparse.ArgumentParser(
        description="Create a Luban v4.x compliant xlsx data file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --output Datas/item.xlsx --sheet TbItem --fields "id:int:ID,name:string:名称,price:int:价格"
  %(prog)s -o Datas/quest.xlsx -s TbQuest -f "id:int:ID,title:string:标题,reward:int:奖励"
        """,
    )
    parser.add_argument("-o", "--output", required=True, help="Output xlsx file path")
    parser.add_argument("-s", "--sheet", required=True, help="Sheet name (typically same as Table name, e.g. TbItem)")
    parser.add_argument(
        "-f", "--fields", required=True,
        help="Field definitions as 'name:type:comment,...' (e.g. 'id:int:ID,name:string:名称')"
    )

    args = parser.parse_args()
    fields = parse_fields(args.fields)

    if not fields:
        print("ERROR: No valid fields provided.", file=sys.stderr)
        sys.exit(1)

    create_xlsx(args.output, args.sheet, fields)


if __name__ == "__main__":
    main()
